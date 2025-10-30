"""
Core Excel workbook processing functionality.
Handles reading, writing, and manipulating Excel workbooks using openpyxl.
"""

import json
from pathlib import Path
from typing import Any, Dict, List, Optional, Union, Tuple
import logging

import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from openpyxl.styles import Font, Fill, Border, Alignment, PatternFill
from openpyxl.formatting.rule import Rule
from openpyxl.worksheet.datavalidation import DataValidation

from ..errors import WorkbookError, WorksheetError, ValidationError
from ..safety import validate_excel_file, validate_sheet_name, validate_cell_reference

logger = logging.getLogger(__name__)


class ExcelProcessor:
    """Core Excel workbook processor with read/write capabilities."""

    def __init__(self):
        self._workbook: Optional[Workbook] = None
        self._file_path: Optional[Path] = None

    def load_workbook(self, file_path: str, read_only: bool = False) -> Dict[str, Any]:
        """
        Load an Excel workbook from file.

        Args:
            file_path: Path to Excel file
            read_only: Whether to open in read-only mode

        Returns:
            Workbook metadata

        Raises:
            WorkbookError: If loading fails
        """
        try:
            validated_path = validate_excel_file(file_path)
            self._file_path = validated_path

            # Load workbook with openpyxl
            self._workbook = openpyxl.load_workbook(
                validated_path,
                read_only=read_only,
                data_only=False,  # Keep formulas
            )

            logger.info(f"Loaded workbook: {validated_path}")

            return self.get_workbook_info()

        except Exception as e:
            raise WorkbookError(f"Failed to load workbook: {e}")

    def get_workbook_info(self) -> Dict[str, Any]:
        """
        Get metadata about the loaded workbook.

        Returns:
            Dictionary with workbook information

        Raises:
            WorkbookError: If no workbook is loaded
        """
        if not self._workbook:
            raise WorkbookError("No workbook loaded")

        try:
            sheet_info = []
            for sheet_name in self._workbook.sheetnames:
                sheet = self._workbook[sheet_name]
                info = {
                    "name": sheet_name,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "data_only": hasattr(sheet, "parent") and sheet.parent.data_only,
                }

                # Get sheet dimensions with data
                if sheet.max_row > 0 and sheet.max_column > 0:
                    info["dimensions"] = (
                        f"A1:{sheet.cell(sheet.max_row, sheet.max_column).coordinate}"
                    )
                else:
                    info["dimensions"] = "A1:A1"

                sheet_info.append(info)

            return {
                "file_path": str(self._file_path) if self._file_path else None,
                "sheet_count": len(self._workbook.sheetnames),
                "sheet_names": self._workbook.sheetnames,
                "sheets": sheet_info,
                "active_sheet": self._workbook.active.title
                if self._workbook.active
                else None,
                "has_formulas": not getattr(self._workbook, "data_only", False),
            }

        except Exception as e:
            raise WorkbookError(f"Failed to get workbook info: {e}")

    def get_worksheet_data(
        self,
        sheet_name: Optional[str] = None,
        include_formulas: bool = False,
        cell_range: Optional[str] = None,
    ) -> Dict[str, Any]:
        """
        Get data from a specific worksheet.

        Args:
            sheet_name: Name of sheet (active sheet if None)
            include_formulas: Whether to include formula strings
            cell_range: Specific cell range to read (e.g., "A1:D10")

        Returns:
            Dictionary with worksheet data

        Raises:
            WorksheetError: If sheet access fails
        """
        if not self._workbook:
            raise WorksheetError("No workbook loaded")

        try:
            # Get worksheet
            if sheet_name:
                validate_sheet_name(sheet_name)
                if sheet_name not in self._workbook.sheetnames:
                    raise WorksheetError(f"Sheet '{sheet_name}' not found")
                sheet = self._workbook[sheet_name]
            else:
                sheet = self._workbook.active
                sheet_name = sheet.title

            # Determine cell range
            if cell_range:
                validate_cell_reference(cell_range)
                cells = sheet[cell_range]
            else:
                # Get all data range
                if sheet.max_row > 0 and sheet.max_column > 0:
                    cells = sheet[
                        f"A1:{sheet.cell(sheet.max_row, sheet.max_column).coordinate}"
                    ]
                else:
                    cells = []

            # Extract cell data
            rows_data = []
            if cells:
                # Handle single cell vs range
                if isinstance(cells, Cell):
                    cells = [[cells]]
                elif len(cells) > 0 and isinstance(cells[0], Cell):
                    cells = [cells]  # Single row

                for row in cells:
                    row_data = []
                    for cell in row:
                        # Skip EmptyCell objects that have no meaningful data
                        if (
                            hasattr(cell, "__class__")
                            and "EmptyCell" in cell.__class__.__name__
                        ):
                            continue

                        cell_info = {
                            "coordinate": cell.coordinate,
                            "value": cell.value,
                            "data_type": cell.data_type,
                            "number_format": cell.number_format,
                        }

                        if include_formulas and cell.data_type == "f":
                            cell_info["formula"] = cell.formula

                        # Add formatting info
                        if cell.font and cell.font != Font():
                            cell_info["font"] = self._serialize_font(cell.font)
                        if cell.fill and cell.fill.patternType:
                            cell_info["fill"] = self._serialize_fill(cell.fill)
                        if cell.alignment and cell.alignment != Alignment():
                            cell_info["alignment"] = self._serialize_alignment(
                                cell.alignment
                            )

                        row_data.append(cell_info)
                    rows_data.append(row_data)

            return {
                "sheet_name": sheet_name,
                "range": cell_range
                or f"A1:{sheet.cell(sheet.max_row, sheet.max_column).coordinate}",
                "rows": len(rows_data),
                "columns": len(rows_data[0]) if rows_data else 0,
                "data": rows_data,
            }

        except Exception as e:
            raise WorksheetError(f"Failed to read worksheet data: {e}")

    def update_cell_value(
        self, sheet_name: str, cell_ref: str, value: Any, formula: Optional[str] = None
    ) -> Dict[str, Any]:
        """
        Update a single cell's value or formula.

        Args:
            sheet_name: Name of the worksheet
            cell_ref: Cell reference (e.g., "A1")
            value: New cell value
            formula: Formula string (if setting a formula)

        Returns:
            Updated cell information

        Raises:
            WorksheetError: If update fails
        """
        if not self._workbook:
            raise WorksheetError("No workbook loaded")

        try:
            validate_sheet_name(sheet_name)
            validate_cell_reference(cell_ref)

            if sheet_name not in self._workbook.sheetnames:
                raise WorksheetError(f"Sheet '{sheet_name}' not found")

            sheet = self._workbook[sheet_name]
            cell = sheet[cell_ref]

            if formula:
                cell.value = f"={formula.lstrip('=')}"
            else:
                cell.value = value

            return {
                "coordinate": cell.coordinate,
                "value": cell.value,
                "data_type": cell.data_type,
                "updated": True,
            }

        except Exception as e:
            raise WorksheetError(f"Failed to update cell: {e}")

    def update_cell_range(
        self, sheet_name: str, cell_range: str, values: List[List[Any]]
    ) -> Dict[str, Any]:
        """
        Update multiple cells in a range.

        Args:
            sheet_name: Name of the worksheet
            cell_range: Cell range (e.g., "A1:C3")
            values: 2D list of values matching the range dimensions

        Returns:
            Update summary

        Raises:
            WorksheetError: If update fails
        """
        if not self._workbook:
            raise WorksheetError("No workbook loaded")

        try:
            validate_sheet_name(sheet_name)
            validate_cell_reference(cell_range)

            if sheet_name not in self._workbook.sheetnames:
                raise WorksheetError(f"Sheet '{sheet_name}' not found")

            sheet = self._workbook[sheet_name]
            cells = sheet[cell_range]

            # Ensure cells is 2D
            if isinstance(cells, Cell):
                cells = [[cells]]
            elif len(cells) > 0 and isinstance(cells[0], Cell):
                cells = [cells]

            updated_count = 0
            for row_idx, cell_row in enumerate(cells):
                if row_idx >= len(values):
                    break
                for col_idx, cell in enumerate(cell_row):
                    if col_idx >= len(values[row_idx]):
                        break
                    cell.value = values[row_idx][col_idx]
                    updated_count += 1

            return {
                "range": cell_range,
                "cells_updated": updated_count,
                "rows_affected": len(cells),
                "columns_affected": len(cells[0]) if cells else 0,
            }

        except Exception as e:
            raise WorksheetError(f"Failed to update cell range: {e}")

    def add_worksheet(
        self, sheet_name: str, index: Optional[int] = None
    ) -> Dict[str, Any]:
        """
        Add a new worksheet to the workbook.

        Args:
            sheet_name: Name for the new worksheet
            index: Position to insert sheet (None for end)

        Returns:
            Information about the created sheet

        Raises:
            WorksheetError: If creation fails
        """
        if not self._workbook:
            raise WorksheetError("No workbook loaded")

        try:
            validated_name = validate_sheet_name(sheet_name)

            if validated_name in self._workbook.sheetnames:
                raise WorksheetError(f"Sheet '{validated_name}' already exists")

            sheet = self._workbook.create_sheet(validated_name, index)

            return {
                "name": sheet.title,
                "index": self._workbook.index(sheet),
                "total_sheets": len(self._workbook.sheetnames),
            }

        except Exception as e:
            raise WorksheetError(f"Failed to add worksheet: {e}")

    def delete_worksheet(self, sheet_name: str) -> Dict[str, Any]:
        """
        Delete a worksheet from the workbook.

        Args:
            sheet_name: Name of sheet to delete

        Returns:
            Deletion summary

        Raises:
            WorksheetError: If deletion fails
        """
        if not self._workbook:
            raise WorksheetError("No workbook loaded")

        try:
            validate_sheet_name(sheet_name)

            if sheet_name not in self._workbook.sheetnames:
                raise WorksheetError(f"Sheet '{sheet_name}' not found")

            if len(self._workbook.sheetnames) <= 1:
                raise WorksheetError("Cannot delete the last remaining worksheet")

            sheet = self._workbook[sheet_name]
            self._workbook.remove(sheet)

            return {
                "deleted_sheet": sheet_name,
                "remaining_sheets": self._workbook.sheetnames,
                "total_sheets": len(self._workbook.sheetnames),
            }

        except Exception as e:
            raise WorksheetError(f"Failed to delete worksheet: {e}")

    def save_workbook(self, file_path: Optional[str] = None) -> Dict[str, Any]:
        """
        Save the workbook to file.

        Args:
            file_path: Path to save to (original path if None)

        Returns:
            Save summary

        Raises:
            WorkbookError: If save fails
        """
        if not self._workbook:
            raise WorkbookError("No workbook loaded")

        try:
            if file_path:
                save_path = Path(file_path)
            elif self._file_path:
                save_path = self._file_path
            else:
                raise WorkbookError("No file path specified for save")

            self._workbook.save(save_path)

            return {
                "saved_to": str(save_path),
                "sheet_count": len(self._workbook.sheetnames),
                "file_size": save_path.stat().st_size if save_path.exists() else 0,
            }

        except Exception as e:
            raise WorkbookError(f"Failed to save workbook: {e}")

    def close_workbook(self) -> None:
        """Close the current workbook and free resources."""
        if self._workbook:
            self._workbook.close()
            self._workbook = None
            self._file_path = None
            logger.info("Workbook closed")

    # Helper methods for serialization
    def _serialize_font(self, font: Font) -> Dict[str, Any]:
        """Serialize font object to dictionary."""
        return {
            "name": font.name,
            "size": font.size,
            "bold": font.bold,
            "italic": font.italic,
            "underline": font.underline,
            "color": str(font.color.rgb) if font.color and font.color.rgb else None,
        }

    def _serialize_fill(self, fill: Fill) -> Dict[str, Any]:
        """Serialize fill object to dictionary."""
        if hasattr(fill, "fgColor") and fill.fgColor:
            return {
                "pattern_type": fill.patternType,
                "fg_color": str(fill.fgColor.rgb) if fill.fgColor.rgb else None,
                "bg_color": str(fill.bgColor.rgb)
                if hasattr(fill, "bgColor") and fill.bgColor and fill.bgColor.rgb
                else None,
            }
        return {"pattern_type": fill.patternType}

    def _serialize_alignment(self, alignment: Alignment) -> Dict[str, Any]:
        """Serialize alignment object to dictionary."""
        return {
            "horizontal": alignment.horizontal,
            "vertical": alignment.vertical,
            "wrap_text": alignment.wrap_text,
            "shrink_to_fit": alignment.shrink_to_fit,
            "indent": alignment.indent,
            "text_rotation": alignment.text_rotation,
        }
