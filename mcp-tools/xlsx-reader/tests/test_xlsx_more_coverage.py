from __future__ import annotations

import json
import shutil
from pathlib import Path

import openpyxl
import pytest

from xlsx_reader import errors, safety
from xlsx_reader.errors import (
    ChartError,
    PivotTableError,
    ValidationError,
    WorkbookError,
    WorksheetError,
)
from xlsx_reader.processors.charts import ChartProcessor
from xlsx_reader.processors.exporters import DataExporter
from xlsx_reader.processors.pivots import PivotTableProcessor
from xlsx_reader.processors.workbook import ExcelProcessor
from xlsx_reader.utils import validation as v


def _create_workbook(path: Path, with_data: bool = True) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"
    wb.create_sheet("Empty")

    if with_data:
        ws["A1"].value = "Product"
        ws["B1"].value = "Qty"
        ws["C1"].value = "Price"
        ws["D1"].value = "Total"

        ws["A2"].value = "Widget"
        ws["B2"].value = 2
        ws["C2"].value = 3
        ws["D2"].value = "=B2*C2"

    wb.save(path)
    wb.close()


def test_errors_helpers_and_extras_cover_branches():
    assert errors.user_input_error("m")["code"] == "UserInput"
    assert errors.user_input_error("m", hint="h")["hint"] == "h"
    assert errors.user_input_error("m", extra=1)["extra"] == 1

    assert errors.forbidden_error("m", x=1)["code"] == "Forbidden"
    assert errors.not_found_error("m", x=1)["code"] == "NotFound"
    assert errors.timeout_error("m", x=1)["code"] == "Timeout"
    assert errors.cancellation_error("m", x=1)["code"] == "Cancelled"

    i1 = errors.internal_error("m", detail="x" * 1000, x=1)
    assert i1["code"] == "Internal"
    assert len(i1["detail"]) == 160

    i2 = errors.internal_error("m", detail=None, x=1)
    assert "detail" not in i2

    ok = errors.success_response({"a": 1}, meta=2)
    assert ok["ok"] is True
    assert ok["meta"] == 2


def test_tools_package_is_importable():
    import xlsx_reader.tools as _t

    assert _t.__all__ == []


def test_validation_more_branches():
    with pytest.raises(ValidationError):
        v.validate_string_param(None, "s", required=True)
    assert v.validate_string_param(None, "s", required=False) is None
    with pytest.raises(ValidationError):
        v.validate_string_param(123, "s")
    with pytest.raises(ValidationError):
        v.validate_string_param("x", "s", min_length=2)
    with pytest.raises(ValidationError):
        v.validate_string_param("xxx", "s", max_length=2)

    with pytest.raises(ValidationError):
        v.validate_int_param(None, "i", required=True)
    assert v.validate_int_param(None, "i", required=False, default=7) == 7
    with pytest.raises(ValidationError):
        v.validate_int_param(True, "i")
    with pytest.raises(ValidationError):
        v.validate_int_param("nope", "i")
    with pytest.raises(ValidationError):
        v.validate_int_param(1, "i", min_value=2)
    with pytest.raises(ValidationError):
        v.validate_int_param(10, "i", max_value=5)

    with pytest.raises(ValidationError):
        v.validate_bool_param(None, "b", required=True)
    assert v.validate_bool_param(None, "b", required=False, default=True) is True
    assert v.validate_bool_param("false", "b") is False
    with pytest.raises(ValidationError):
        v.validate_bool_param("maybe", "b")

    with pytest.raises(ValidationError):
        v.validate_list_param(None, "l", required=True)
    assert v.validate_list_param(None, "l", required=False) is None
    with pytest.raises(ValidationError):
        v.validate_list_param("x", "l")
    with pytest.raises(ValidationError):
        v.validate_list_param([], "l", min_length=1)
    with pytest.raises(ValidationError):
        v.validate_list_param(["a", "b"], "l", max_length=1)
    with pytest.raises(ValidationError):
        v.validate_list_param(["a", 2], "l", item_type=str)

    with pytest.raises(ValidationError):
        v.validate_choice_param(None, "c", ["a"], required=True)
    assert v.validate_choice_param(None, "c", ["a"], required=False, default="a") == "a"
    with pytest.raises(ValidationError):
        v.validate_choice_param("x", "c", ["a", "b"])  # invalid

    with pytest.raises(ValidationError):
        v.validate_dict_param(None, "d", required=True)
    assert v.validate_dict_param(None, "d", required=False) is None
    with pytest.raises(ValidationError):
        v.validate_dict_param("x", "d")
    with pytest.raises(ValidationError):
        v.validate_dict_param({"a": 1}, "d", required_keys={"b"})
    with pytest.raises(ValidationError):
        v.validate_dict_param({"a": 1, "x": 2}, "d", allowed_keys={"a"})


def test_safety_more_branches(tmp_path: Path, monkeypatch):
    # validate_file_path bad inputs
    with pytest.raises(ValidationError):
        safety.validate_file_path("")

    class BadPath:
        def resolve(self):
            raise RuntimeError("boom")

    monkeypatch.setattr(safety, "Path", lambda _p: BadPath())
    with pytest.raises(ValidationError):
        safety.validate_file_path("/nope")

    # restore Path
    monkeypatch.setattr(safety, "Path", Path)

    # missing
    with pytest.raises(errors.FileAccessError):
        safety.validate_file_path(str(tmp_path / "missing.xlsx"))

    # not a file
    d = tmp_path / "dir.xlsx"
    d.mkdir()
    with pytest.raises(errors.FileAccessError):
        safety.validate_file_path(str(d))

    # extension invalid
    bad = tmp_path / "a.txt"
    bad.write_text("x", encoding="utf-8")
    with pytest.raises(ValidationError):
        safety.validate_excel_file(str(bad))

    # validate_file_size too large and stat error
    xlsx = tmp_path / "book.xlsx"
    _create_workbook(xlsx)

    class FakeStat:
        st_size = safety.MAX_FILE_SIZE_BYTES + 1

    monkeypatch.setattr(Path, "stat", lambda self: FakeStat())
    with pytest.raises(ValidationError):
        safety.validate_file_size(xlsx)

    def boom_stat(self):
        raise OSError("no")

    monkeypatch.setattr(Path, "stat", boom_stat)
    with pytest.raises(errors.FileAccessError):
        safety.validate_file_size(xlsx)

    # restore Path.stat
    monkeypatch.undo()

    # create_backup failure
    def boom_copy(*args, **kwargs):
        raise OSError("no")

    monkeypatch.setattr(shutil, "copy2", boom_copy)
    with pytest.raises(errors.FileAccessError):
        safety.create_backup(xlsx)


def test_safety_restore_and_cleanup_edge_cases(tmp_path: Path, monkeypatch):
    xlsx = tmp_path / "book.xlsx"
    _create_workbook(xlsx)

    missing_backup = tmp_path / "book.xlsx.backup"
    with pytest.raises(errors.FileAccessError):
        safety.restore_backup(xlsx, missing_backup)

    # restore failure
    backup = tmp_path / "book.xlsx.backup"
    backup.write_text("x", encoding="utf-8")

    def boom_copy(*args, **kwargs):
        raise OSError("no")

    monkeypatch.setattr(shutil, "copy2", boom_copy)
    with pytest.raises(errors.FileAccessError):
        safety.restore_backup(xlsx, backup)

    # cleanup failure branch (logs warning)
    def boom_unlink(self):
        raise OSError("no")

    monkeypatch.setattr(Path, "unlink", boom_unlink)
    safety.cleanup_backup(backup)


def test_sheet_and_cell_validation_more():
    with pytest.raises(ValidationError):
        safety.validate_sheet_name("")
    with pytest.raises(ValidationError):
        safety.validate_sheet_name("x" * 40)
    with pytest.raises(ValidationError):
        safety.validate_sheet_name("bad/")
    assert safety.validate_sheet_name("  OK ") == "OK"

    with pytest.raises(ValidationError):
        safety.validate_cell_reference("")
    with pytest.raises(ValidationError):
        safety.validate_cell_reference("A1:B2:C3")
    with pytest.raises(ValidationError):
        safety.validate_cell_reference("AA")
    assert safety.validate_cell_reference("a1") == "A1"


def test_excel_processor_more_errors_and_formatting(tmp_path: Path):
    xlsx = tmp_path / "book.xlsx"
    _create_workbook(xlsx)

    proc = ExcelProcessor()

    # No workbook loaded paths
    with pytest.raises(WorksheetError):
        proc.get_worksheet_data("Sales")
    with pytest.raises(WorkbookError):
        proc.save_workbook()

    proc.load_workbook(str(xlsx), read_only=False)

    # Sheet not found
    with pytest.raises(WorksheetError):
        proc.get_worksheet_data("Missing")

    # Invalid cell range
    with pytest.raises(WorksheetError):
        proc.get_worksheet_data("Sales", cell_range="A1:B2:C3")

    # update errors
    with pytest.raises(WorksheetError):
        proc.update_cell_value("Missing", "A1", 1)

    # add duplicate sheet
    with pytest.raises(WorksheetError):
        proc.add_worksheet("Sales")

    # delete last remaining worksheet
    # Create a single-sheet workbook
    single = tmp_path / "single.xlsx"
    wb = openpyxl.Workbook()
    wb.save(single)
    wb.close()

    proc2 = ExcelProcessor()
    proc2.load_workbook(str(single), read_only=False)
    with pytest.raises(WorksheetError):
        proc2.delete_worksheet(proc2._workbook.sheetnames[0])


def test_exporter_more_paths(tmp_path: Path, monkeypatch):
    xlsx = tmp_path / "book.xlsx"
    _create_workbook(xlsx)

    proc = ExcelProcessor()
    proc.load_workbook(str(xlsx), read_only=False)
    exporter = DataExporter(proc)

    # Empty sheet export -> success_response path
    orig_get = proc.get_worksheet_data
    monkeypatch.setattr(proc, "get_worksheet_data", lambda _name: {"data": []})
    r = exporter.export_worksheet_to_csv("Empty")
    assert r["ok"] is True

    monkeypatch.setattr(proc, "get_worksheet_data", orig_get)

    # Save CSV to file
    out_csv = tmp_path / "out.csv"
    r2 = exporter.export_worksheet_to_csv("Sales", output_path=str(out_csv), delimiter=";")
    assert Path(r2["saved_to"]).exists()

    # JSON export to file
    out_json = tmp_path / "out.json"
    r3 = exporter.export_workbook_to_json(output_path=str(out_json), include_formulas=True, include_formatting=False)
    assert Path(r3["saved_to"]).exists()

    # Sheet export failure branch during JSON export
    def boom_get(sheet_name=None, **_kwargs):
        if sheet_name == "Sales":
            raise WorksheetError("boom")
        return {"range": "A1:A1", "rows": 0, "columns": 0, "data": []}

    monkeypatch.setattr(proc, "get_worksheet_data", boom_get)
    j = exporter.export_workbook_to_json()
    assert "json_data" in j

    # pandas empty
    df_empty = exporter.export_sheet_to_pandas("Empty")
    assert df_empty.empty is True

    # include_headers False branch
    monkeypatch.setattr(proc, "get_worksheet_data", lambda name: {"data": [[{"value": 1}], [{"value": 2}]]})
    df = exporter.export_sheet_to_pandas("Sales", include_headers=False)
    assert df.shape[0] == 2


def test_chart_processor_more_error_paths(tmp_path: Path):
    xlsx = tmp_path / "book.xlsx"
    _create_workbook(xlsx)

    wb = openpyxl.load_workbook(xlsx, read_only=False, data_only=False)
    charts = ChartProcessor(wb)

    with pytest.raises(ChartError):
        charts.extract_charts_from_sheet("Missing")

    # Unsupported type
    with pytest.raises(ChartError):
        charts.create_chart("Sales", "nope", "Sales!A1:D2")

    # modify/delete without charts
    with pytest.raises(ChartError):
        charts.modify_chart("Sales", 0, {"title": "x"})
    with pytest.raises(ChartError):
        charts.delete_chart("Sales", 0)

    # extract_all_charts includes per-sheet error dict
    all_charts = charts.extract_all_charts()
    assert "Sales" not in all_charts or isinstance(all_charts.get("Sales"), (list, dict))

    wb.close()


def test_pivot_processor_error_paths(tmp_path: Path):
    xlsx = tmp_path / "book.xlsx"
    _create_workbook(xlsx)

    wb = openpyxl.load_workbook(xlsx, read_only=False, data_only=False)
    piv = PivotTableProcessor(wb)

    with pytest.raises(PivotTableError):
        piv.extract_pivot_tables_from_sheet("Missing")

    with pytest.raises(PivotTableError):
        piv.get_pivot_data_summary("Sales", 0)

    # create/modify/delete errors on missing pivots
    with pytest.raises(PivotTableError):
        piv.modify_pivot_table("Sales", 0, {})

    with pytest.raises(PivotTableError):
        piv.delete_pivot_table("Sales", 0)

    # create_pivot_table missing sheets
    with pytest.raises(PivotTableError):
        piv.create_pivot_table("Nope", "A1:A1", "Sales", "A1", {})

    with pytest.raises(PivotTableError):
        piv.create_pivot_table("Sales", "A1:A1", "Nope", "A1", {})

    wb.close()
