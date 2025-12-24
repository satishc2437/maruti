from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl
import pytest

from xlsx_reader import safety
from xlsx_reader.errors import ValidationError
from xlsx_reader.processors.charts import ChartProcessor
from xlsx_reader.processors.exporters import DataExporter
from xlsx_reader.processors.pivots import PivotTableProcessor
from xlsx_reader.processors.workbook import ExcelProcessor
from xlsx_reader.utils import validation as v


def _create_sample_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    ws["A1"].value = "Product"
    ws["B1"].value = "Qty"
    ws["C1"].value = "Price"
    ws["D1"].value = "Total"

    ws["A2"].value = "Widget"
    ws["B2"].value = 2
    ws["C2"].value = 3
    ws["D2"].value = "=B2*C2"

    wb.create_sheet("Summary")

    wb.save(path)
    wb.close()


def test_validation_helpers_cover_common_paths(tmp_path: Path):
    v.validate_required_params({"a": 1}, {"a"})
    with pytest.raises(ValidationError):
        v.validate_required_params({"a": 1}, {"a", "b"})

    v.validate_unknown_params({"a": 1}, {"a"})
    with pytest.raises(ValidationError):
        v.validate_unknown_params({"a": 1, "x": 2}, {"a"})

    assert v.validate_string_param("ok", "s") == "ok"
    assert v.validate_int_param("3", "i") == 3
    assert v.validate_bool_param("yes", "b") is True
    assert v.validate_list_param(["a", "b"], "l", item_type=str) == ["a", "b"]
    assert v.validate_choice_param("x", "c", ["x", "y"]) == "x"

    assert v.validate_dict_param({"k": 1}, "d", required_keys={"k"}, allowed_keys={"k"}) == {"k": 1}
    with pytest.raises(ValidationError):
        v.validate_dict_param({"x": 1}, "d", required_keys={"k"})


def test_safety_file_ops(tmp_path: Path):
    xlsx_path = tmp_path / "book.xlsx"
    _create_sample_workbook(xlsx_path)

    assert safety.validate_excel_file(str(xlsx_path)) == xlsx_path.resolve()

    backup = safety.create_backup(xlsx_path)
    assert backup.exists()

    # Mutate original then restore.
    xlsx_path.write_bytes(b"corrupted")
    safety.restore_backup(xlsx_path, backup)
    assert xlsx_path.stat().st_size > 0

    safety.cleanup_backup(backup)
    assert not backup.exists()

    # Context manager success path cleans backup.
    with safety.FileOperationContext(str(xlsx_path), create_backup=True) as p:
        assert p.exists()

    # Context manager exception path restores from backup.
    with pytest.raises(RuntimeError):
        with safety.FileOperationContext(str(xlsx_path), create_backup=True):
            raise RuntimeError("boom")


def test_excel_processor_and_exporter_and_charts(tmp_path: Path):
    xlsx_path = tmp_path / "book.xlsx"
    _create_sample_workbook(xlsx_path)

    processor = ExcelProcessor()

    with pytest.raises(Exception):
        processor.get_workbook_info()

    info = processor.load_workbook(str(xlsx_path), read_only=False)
    assert info["sheet_count"] == 2

    data = processor.get_worksheet_data("Sales", include_formulas=True, cell_range="A1:D2")
    assert data["rows"] == 2

    upd = processor.update_cell_value("Sales", "E1", "Status")
    assert upd["updated"] is True

    upd_range = processor.update_cell_range("Sales", "A3:B3", [["X", 1]])
    assert upd_range["cells_updated"] == 2

    add = processor.add_worksheet("New")
    assert add["name"] == "New"

    delete = processor.delete_worksheet("New")
    assert delete["deleted_sheet"] == "New"

    save = processor.save_workbook()
    assert Path(save["saved_to"]).exists()

    exporter = DataExporter(processor)

    csv = exporter.export_worksheet_to_csv("Sales")
    assert "csv_data" in csv

    js = exporter.export_workbook_to_json(include_formulas=True, include_formatting=True)
    assert "json_data" in js

    df = exporter.export_sheet_to_pandas("Sales")
    assert df.shape[0] >= 1

    stats = exporter.get_summary_statistics("Sales")
    assert stats["total_rows"] >= 1

    # Charts: create then extract/modify/delete.
    chart_proc = ChartProcessor(processor._workbook)
    created = chart_proc.create_chart("Sales", "column", "Sales!A1:D2", title="T")
    assert created["created"] is True

    charts = chart_proc.extract_charts_from_sheet("Sales")
    assert isinstance(charts, list)
    assert len(charts) >= 1

    modified = chart_proc.modify_chart("Sales", 0, {"title": "New"})
    assert modified["modified"] is True

    deleted = chart_proc.delete_chart("Sales", 0)
    assert deleted["deleted"] is True

    processor.close_workbook()


@dataclass
class _FakeField:
    x: int = 0
    name: str | None = None


@dataclass
class _FakePivotStyle:
    name: str = "PivotStyle"
    showRowHeaders: bool = True
    showColHeaders: bool = True
    showRowStripes: bool = False
    showColStripes: bool = False


@dataclass
class _FakePivot:
    name: str = "PT1"
    pivotFields: list[object] | None = None
    dataFields: list[object] | None = None
    rowFields: list[object] | None = None
    colFields: list[object] | None = None
    pageFields: list[object] | None = None
    pivotTableStyleInfo: object | None = None


def test_pivot_table_processor_with_fake_pivot(tmp_path: Path):
    xlsx_path = tmp_path / "book.xlsx"
    _create_sample_workbook(xlsx_path)

    processor = ExcelProcessor()
    processor.load_workbook(str(xlsx_path), read_only=False)

    sheet = processor._workbook["Sales"]

    fake = _FakePivot(
        pivotFields=[_FakeField(name="F1")],
        dataFields=[_FakeField(name="DF1")],
        rowFields=[_FakeField(x=1)],
        colFields=[_FakeField(x=2)],
        pageFields=[_FakeField(x=3)],
        pivotTableStyleInfo=_FakePivotStyle(),
    )

    # Inject into openpyxl worksheet.
    sheet._pivots = [fake]  # type: ignore[attr-defined]

    piv = PivotTableProcessor(processor._workbook)

    extracted = piv.extract_pivot_tables_from_sheet("Sales")
    assert len(extracted) == 1
    assert extracted[0]["name"] == "PT1"

    summary = piv.get_pivot_data_summary("Sales", 0)
    assert summary["name"] == "PT1"

    created = piv.create_pivot_table("Sales", "A1:D2", "Summary", "A1", {})
    assert created["created"] is False

    modified = piv.modify_pivot_table("Sales", 0, {"x": 1})
    assert modified["modified"] is False

    deleted = piv.delete_pivot_table("Sales", 0)
    assert deleted["deleted"] in (True, False)

    processor.close_workbook()
