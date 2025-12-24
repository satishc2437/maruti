from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import openpyxl
import pytest
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlWorksheet

from xlsx_reader.errors import ChartError, PivotTableError, WorkbookError, WorksheetError
from xlsx_reader.processors.charts import ChartProcessor
from xlsx_reader.processors.exporters import DataExporter
from xlsx_reader.processors.pivots import PivotTableProcessor
from xlsx_reader.processors.workbook import ExcelProcessor
from xlsx_reader.safety import FileOperationContext, cleanup_backup, restore_backup
from xlsx_reader.utils.validation import validate_bool_param


def _create_workbook(path: Path) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales"

    ws["A1"].value = "Name"
    ws["B1"].value = "Qty"
    ws["C1"].value = "Note"

    ws["A2"].value = "Widget"
    ws["B2"].value = 2
    ws["C2"].value = None

    # Some formatting + formula coverage
    ws["A1"].font = Font(bold=True)
    ws["A1"].alignment = Alignment(wrap_text=True)
    ws["A1"].fill = PatternFill(patternType="solid", fgColor="FF0000")
    ws["B3"].value = "=SUM(B2:B2)"

    wb.create_sheet("Other")

    wb.save(path)
    wb.close()


def test_chart_processor_more_branches(tmp_path: Path, monkeypatch):
    xlsx = tmp_path / "book.xlsx"
    _create_workbook(xlsx)

    wb = openpyxl.load_workbook(xlsx, read_only=False, data_only=False)
    charts = ChartProcessor(wb)

    created = charts.create_chart(
        "Sales",
        "column",
        "Sales!A1:B2",
        title="T",
        position={"col": 2, "row": 10},
    )
    assert created["created"] is True

    extracted = charts.extract_charts_from_sheet("Sales")
    assert extracted and extracted[0]["position"]["anchor"]

    # Modify: cover title/data_range/position/unknown property + failure to apply
    modified = charts.modify_chart(
        "Sales",
        0,
        {
            "title": "New",
            "data_range": "Sales!A1:B2",
            "position": {"col": 3, "row": 3},
            "unknown": 1,
        },
    )
    assert modified["modified"] is True

    # Trigger a failure applying a real modification key
    charts.modify_chart("Sales", 0, {"data_range": "NOT_A_RANGE"})

    with pytest.raises(ChartError):
        charts.modify_chart("Sales", 99, {"title": "x"})

    with pytest.raises(ChartError):
        charts.delete_chart("Sales", 99)

    deleted = charts.delete_chart("Sales", 0)
    assert deleted["deleted"] is True

    # extract_all_charts exception path for a sheet
    def boom(_sheet_name: str):
        raise RuntimeError("boom")

    monkeypatch.setattr(charts, "extract_charts_from_sheet", boom)
    all_charts = charts.extract_all_charts()
    assert any(isinstance(v, dict) and "error" in v for v in all_charts.values())

    wb.close()


def test_chart_helper_exception_fallbacks(tmp_path: Path):
    xlsx = tmp_path / "book.xlsx"
    _create_workbook(xlsx)
    wb = openpyxl.load_workbook(xlsx, read_only=False, data_only=False)
    charts = ChartProcessor(wb)

    class BadAnchor:
        @property
        def col(self):
            raise RuntimeError("boom")

        @property
        def row(self):
            raise RuntimeError("boom")

        def __str__(self):
            raise RuntimeError("boom")

    class BadChart:
        anchor = BadAnchor()

        @property
        def series(self):
            raise RuntimeError("boom")

        @property
        def legend(self):
            raise RuntimeError("boom")

        @property
        def style(self):
            raise RuntimeError("boom")

    assert charts._get_chart_position(BadChart())["anchor"] is None
    assert charts._get_chart_data_range(BadChart()) is None
    assert charts._extract_chart_series(BadChart()) == []
    assert charts._get_chart_style(BadChart()) == {}

    wb.close()


def test_chart_processor_unsupported_type(tmp_path: Path):
    xlsx = tmp_path / "book.xlsx"
    _create_workbook(xlsx)
    wb = openpyxl.load_workbook(xlsx, read_only=False, data_only=False)
    charts = ChartProcessor(wb)
    with pytest.raises(ChartError):
        charts.create_chart("Sales", "nope", "Sales!A1:B2")
    wb.close()


@dataclass
class _FakeCache:
    worksheetSource: str = "A1:B2"
    recordCount: int = 2
    refreshOnLoad: bool = True


@dataclass
class _FakeLocation:
    ref: str = "A1"
    firstHeaderRow: int = 1
    firstDataRow: int = 2
    firstDataCol: int = 1


@dataclass
class _FakeField:
    name: str | None = None
    axis: str | None = None
    dataField: bool = False
    items: list[object] | None = None
    x: int = 0


@dataclass
class _FakeStyle:
    name: str = "PivotStyle"
    showRowHeaders: bool = True
    showColHeaders: bool = True
    showRowStripes: bool = False
    showColStripes: bool = False


@dataclass
class _FakePivot:
    name: str = "PT1"
    cache: _FakeCache | None = None
    location: _FakeLocation | None = None
    pivotFields: list[_FakeField] | None = None
    dataFields: list[_FakeField] | None = None
    rowFields: list[_FakeField] | None = None
    colFields: list[_FakeField] | None = None
    pageFields: list[_FakeField] | None = None
    pivotTableStyleInfo: _FakeStyle | None = None


def test_pivot_processor_more_branches(tmp_path: Path, monkeypatch):
    xlsx = tmp_path / "book.xlsx"
    _create_workbook(xlsx)

    wb = openpyxl.load_workbook(xlsx, read_only=False, data_only=False)
    sheet = wb["Sales"]

    # Inject richer pivot
    sheet._pivots = [
        _FakePivot(
            cache=_FakeCache(),
            location=_FakeLocation(),
            pivotFields=[_FakeField(name="F1", axis="row", items=[1])],
            dataFields=[_FakeField(name="DF1", axis=None)],
            rowFields=[_FakeField(x=1)],
            colFields=[_FakeField(x=2)],
            pageFields=[_FakeField(x=3)],
            pivotTableStyleInfo=_FakeStyle(),
        )
    ]  # type: ignore[attr-defined]

    piv = PivotTableProcessor(wb)

    extracted = piv.extract_pivot_tables_from_sheet("Sales")
    assert extracted and extracted[0]["cache_definition"]["record_count"] == 2

    summary = piv.get_pivot_data_summary("Sales", 0)
    assert summary["field_count"] >= 1

    # extract_all_pivot_tables error dict branch
    def boom(_sheet_name: str):
        raise RuntimeError("boom")

    monkeypatch.setattr(piv, "extract_pivot_tables_from_sheet", boom)
    all_pivots = piv.extract_all_pivot_tables()
    assert any(isinstance(v, dict) and "error" in v for v in all_pivots.values())

    # delete_pivot_table inner exception sets deleted False
    class BadList(list):
        def pop(self, idx=-1):
            raise RuntimeError("no")

    sheet._pivots = BadList([object()])  # type: ignore[attr-defined]
    deleted = PivotTableProcessor(wb).delete_pivot_table("Sales", 0)
    assert deleted["deleted"] is False

    wb.close()


def test_pivot_processor_more_errors_and_helpers(tmp_path: Path):
    xlsx = tmp_path / "book.xlsx"
    _create_workbook(xlsx)
    wb = openpyxl.load_workbook(xlsx, read_only=False, data_only=False)
    sheet = wb["Sales"]
    piv = PivotTableProcessor(wb)

    # Normal extract_all path where only some sheets have pivots
    sheet._pivots = [_FakePivot()]  # type: ignore[attr-defined]
    all_pivots = piv.extract_all_pivot_tables()
    assert "Sales" in all_pivots

    # Sheet not found
    with pytest.raises(PivotTableError):
        piv.extract_pivot_tables_from_sheet("Missing")

    # No pivots
    sheet._pivots = []  # type: ignore[attr-defined]
    with pytest.raises(PivotTableError):
        piv.modify_pivot_table("Sales", 0, {"x": 1})

    # Index out of range
    sheet._pivots = [_FakePivot()]  # type: ignore[attr-defined]
    with pytest.raises(PivotTableError):
        piv.modify_pivot_table("Sales", 99, {})
    with pytest.raises(PivotTableError):
        piv.delete_pivot_table("Sales", 99)
    with pytest.raises(PivotTableError):
        piv.get_pivot_data_summary("Sales", 99)

    # create_pivot_table required-fields defaulting
    created = piv.create_pivot_table(
        "Sales",
        "A1:B2",
        "Sales",
        "A10",
        {"row_fields": ["A"]},
    )
    assert "column_fields" in created["fields_config"]
    assert "data_fields" in created["fields_config"]

    # Helper exception fallbacks
    class BadPivot:
        @property
        def cache(self):
            raise RuntimeError("boom")

        @property
        def location(self):
            raise RuntimeError("boom")

        @property
        def pivotFields(self):
            raise RuntimeError("boom")

        @property
        def dataFields(self):
            raise RuntimeError("boom")

        @property
        def rowFields(self):
            raise RuntimeError("boom")

        @property
        def colFields(self):
            raise RuntimeError("boom")

        @property
        def pageFields(self):
            raise RuntimeError("boom")

        @property
        def pivotTableStyleInfo(self):
            raise RuntimeError("boom")

    assert piv._get_cache_definition(BadPivot()) == {}
    assert piv._get_pivot_location(BadPivot()) == {}
    assert piv._extract_pivot_fields(BadPivot()) == []
    assert piv._extract_data_fields(BadPivot()) == []
    assert piv._extract_row_fields(BadPivot()) == []
    assert piv._extract_column_fields(BadPivot()) == []
    assert piv._extract_filter_fields(BadPivot()) == []
    assert piv._get_pivot_style(BadPivot()) == {}

    wb.close()


def test_workbook_processor_more_branches(tmp_path: Path, monkeypatch):
    xlsx = tmp_path / "book.xlsx"
    _create_workbook(xlsx)

    proc = ExcelProcessor()

    # load_workbook error
    monkeypatch.setattr(openpyxl, "load_workbook", lambda *_a, **_k: (_ for _ in ()).throw(RuntimeError("boom")))
    with pytest.raises(Exception):
        proc.load_workbook(str(xlsx), read_only=False)

    # Restore and load
    monkeypatch.undo()
    proc.load_workbook(str(xlsx), read_only=False)

    # get_worksheet_data single-cell range branch
    out = proc.get_worksheet_data("Sales", cell_range="A1", include_formulas=False)
    assert out["rows"] == 1

    # update_cell_range with short values triggers breaks
    upd = proc.update_cell_range("Sales", "A1:C2", [["X"], ["Y", 2]])
    assert upd["cells_updated"] >= 2

    # add at index
    added = proc.add_worksheet("Inserted", index=0)
    assert added["index"] == 0

    # delete not-found
    with pytest.raises(WorksheetError):
        proc.delete_worksheet("Nope")

    # save_workbook file_path override
    out_path = tmp_path / "out.xlsx"
    saved = proc.save_workbook(str(out_path))
    assert Path(saved["saved_to"]).exists()

    proc.close_workbook()


def test_workbook_processor_formatting_and_save_paths(tmp_path: Path):
    xlsx = tmp_path / "book.xlsx"
    _create_workbook(xlsx)

    proc = ExcelProcessor()
    proc.load_workbook(str(xlsx), read_only=False)

    info = proc.get_workbook_info()
    assert info["sheet_count"] >= 1

    # Active sheet branch + include_formulas path + formatting serialization
    data = proc.get_worksheet_data(sheet_name=None, include_formulas=True)
    assert data["sheet_name"]
    assert any(
        any("font" in c or "fill" in c or "alignment" in c for c in row)
        for row in data["data"]
    )
    assert any(
        any(c.get("formula") for c in row if c.get("data_type") == "f")
        for row in data["data"]
    )

    # update_cell_value formula branch
    updated = proc.update_cell_value("Sales", "C3", value=None, formula="1+1")
    assert updated["updated"] is True
    assert str(updated["value"]).startswith("=")

    # save_workbook uses original file path when None
    saved = proc.save_workbook()
    assert Path(saved["saved_to"]).exists()

    # Error branch: no file path specified
    proc2 = ExcelProcessor()
    proc2._workbook = openpyxl.Workbook()  # intentional for coverage
    with pytest.raises(WorkbookError):
        proc2.save_workbook()

    proc.close_workbook()


def test_workbook_processor_remaining_branches(tmp_path: Path, monkeypatch):
    xlsx = tmp_path / "book.xlsx"
    _create_workbook(xlsx)

    proc = ExcelProcessor()
    proc.load_workbook(str(xlsx), read_only=False)

    # get_workbook_info: force dimensions else-branch
    monkeypatch.setattr(OpenpyxlWorksheet, "max_row", property(lambda _self: 0), raising=False)
    monkeypatch.setattr(OpenpyxlWorksheet, "max_column", property(lambda _self: 0), raising=False)
    info = proc.get_workbook_info()
    assert any(s.get("dimensions") == "A1:A1" for s in info.get("sheets", []))

    # get_workbook_info exception handler
    from openpyxl.workbook.workbook import Workbook as OpenpyxlWorkbook

    def boom_getitem(self, _key):
        raise RuntimeError("boom")

    monkeypatch.setattr(OpenpyxlWorkbook, "__getitem__", boom_getitem)
    with pytest.raises(WorkbookError):
        proc.get_workbook_info()

    # Restore monkeypatches before loading a new workbook instance.
    monkeypatch.undo()

    # Restore workbook for worksheet-data tests
    proc2 = ExcelProcessor()
    proc2.load_workbook(str(xlsx), read_only=False)

    # get_worksheet_data: cells=[] branch by forcing an empty read for the computed range
    orig_getitem = OpenpyxlWorksheet.__getitem__

    def empty_range_getitem(self, key):
        if isinstance(key, str) and key.startswith("A1:"):
            return []
        return orig_getitem(self, key)

    monkeypatch.setattr(OpenpyxlWorksheet, "__getitem__", empty_range_getitem)
    data = proc2.get_worksheet_data(sheet_name="Sales", include_formulas=False, cell_range=None)
    assert data["rows"] == 0

    # get_worksheet_data: single-row coercion branch + update_cell_range branches
    monkeypatch.undo()
    one_row = proc2.get_worksheet_data(sheet_name="Sales", cell_range="A1:C1")
    assert one_row["rows"] == 1

    # update_cell_value/update_cell_range: no-workbook-loaded branches
    with pytest.raises(WorksheetError):
        ExcelProcessor().update_cell_value("Sales", "A1", value=1)
    with pytest.raises(WorksheetError):
        ExcelProcessor().update_cell_range("Sales", "A1", [[1]])

    # update_cell_range: sheet not found
    with pytest.raises(WorksheetError):
        proc2.update_cell_range("Missing", "A1", [[1]])

    # update_cell_range: single-cell + single-row coercions + early row break
    proc2.update_cell_range("Sales", "A1", [["Z"]])
    proc2.update_cell_range("Sales", "A1:C1", [["a", "b", "c"]])
    proc2.update_cell_range("Sales", "A1:C2", [["r1", "r1", "r1"]])

    # update_cell_range exception handler
    with pytest.raises(WorksheetError):
        proc2.update_cell_range("Sales", "A1:B2:C3", [[1]])

    # add/delete/close: no-workbook-loaded branches
    with pytest.raises(WorksheetError):
        ExcelProcessor().add_worksheet("X")
    with pytest.raises(WorksheetError):
        ExcelProcessor().delete_worksheet("X")
    ExcelProcessor().close_workbook()

    # delete_worksheet: no-workbook-loaded already, also hit last remaining-sheet by creating 1-sheet workbook
    proc_single = ExcelProcessor()
    wb_single = openpyxl.Workbook()
    proc_single._workbook = wb_single
    with pytest.raises(WorksheetError):
        proc_single.delete_worksheet(wb_single.active.title)

    proc2.close_workbook()


def test_workbook_get_worksheet_data_skip_emptycell_and_formula_attr(tmp_path: Path, monkeypatch):
    xlsx = tmp_path / "book.xlsx"
    _create_workbook(xlsx)
    proc = ExcelProcessor()
    proc.load_workbook(str(xlsx), read_only=False)

    class EmptyCell:
        pass

    class FakeCell:
        coordinate = "B2"
        value = 1
        data_type = "f"
        number_format = "General"
        formula = "=1+1"
        font = Font(bold=True)
        # Use a minimal fill-like object to trigger the _serialize_fill else-branch.
        # We avoid openpyxl's descriptors which require fgColor to be a Color.
        class FillNoFg:
            patternType = "solid"

        fill = FillNoFg()
        alignment = Alignment(wrap_text=True)

    fake_cell = FakeCell()

    orig_getitem = OpenpyxlWorksheet.__getitem__

    def patched_getitem(self, key):
        if key == "A1":
            return [[EmptyCell()]]
        if key == "B2":
            return [[fake_cell]]
        return orig_getitem(self, key)

    monkeypatch.setattr(OpenpyxlWorksheet, "__getitem__", patched_getitem)

    skipped = proc.get_worksheet_data("Sales", cell_range="A1")
    assert skipped["rows"] == 1
    assert skipped["columns"] == 0  # EmptyCell skipped

    out = proc.get_worksheet_data("Sales", cell_range="B2", include_formulas=True)
    assert out["data"][0][0]["formula"] == "=1+1"
    assert out["data"][0][0]["fill"] == {"pattern_type": "solid"}

    proc.close_workbook()


def test_safety_remaining_branches(tmp_path: Path, monkeypatch):
    xlsx = tmp_path / "book.xlsx"
    _create_workbook(xlsx)

    # restore_backup: backup_path defaulting line
    backup_path = xlsx.with_suffix(xlsx.suffix + ".backup")
    backup_path.write_bytes(xlsx.read_bytes())
    restore_backup(xlsx)

    # cleanup_backup: no-op branch when backup doesn't exist
    cleanup_backup(tmp_path / "missing.xlsx.backup")

    # FileOperationContext: create_backup False branch
    with FileOperationContext(str(xlsx), create_backup=False):
        pass

    # FileOperationContext: __enter__ exception path + cleanup
    from xlsx_reader import safety as safety_mod

    monkeypatch.setattr(safety_mod.FileLock, "acquire", lambda *_a, **_k: (_ for _ in ()).throw(RuntimeError("boom")))
    monkeypatch.setattr(safety_mod.FileLock, "release", lambda *_a, **_k: None)
    with pytest.raises(Exception):
        with FileOperationContext(str(xlsx), create_backup=False):
            pass

    # _cleanup branch where file_lock is None (216->exit)
    ctx = FileOperationContext(str(xlsx), create_backup=False)
    ctx.file_lock = None
    ctx._cleanup()

    # _cleanup release failure branch
    class BadLock:
        def release(self):
            raise RuntimeError("no")

    ctx.file_lock = BadLock()  # type: ignore[assignment]
    ctx._cleanup()

    # validate_cell_reference: empty cell part
    from xlsx_reader.safety import validate_cell_reference

    with pytest.raises(Exception):
        validate_cell_reference(":A1")


def test_charts_pivots_exporters_remaining_branches(tmp_path: Path, monkeypatch):
    xlsx = tmp_path / "book.xlsx"
    _create_workbook(xlsx)
    wb = openpyxl.load_workbook(xlsx, read_only=False, data_only=False)
    charts = ChartProcessor(wb)
    pivots = PivotTableProcessor(wb)

    # charts.extract_all_charts normal path (line 76)
    charts.create_chart("Sales", "column", "Sales!A1:B2")
    all_charts = charts.extract_all_charts()
    assert "Sales" in all_charts

    # charts outer exception handler
    class BadWB:
        @property
        def sheetnames(self):
            raise RuntimeError("boom")

    with pytest.raises(ChartError):
        ChartProcessor(BadWB()).extract_all_charts()  # type: ignore[arg-type]

    # chart sheet-not-found branches
    with pytest.raises(ChartError):
        charts.create_chart("Missing", "column", "Sales!A1:B2")
    with pytest.raises(ChartError):
        charts.modify_chart("Missing", 0, {})
    with pytest.raises(ChartError):
        charts.delete_chart("Missing", 0)

    # helper return chart_class + data_range return
    class WeirdChart:
        pass

    assert charts._get_chart_type(WeirdChart()) == "weirdchart"

    class Series:
        values = "Sales!A1:B2"

    class HasSeries:
        series = [Series()]

    assert charts._get_chart_data_range(HasSeries()) == "Sales!A1:B2"

    # pivots outer exception handler
    class BadWB2:
        @property
        def sheetnames(self):
            raise RuntimeError("boom")

    with pytest.raises(PivotTableError):
        PivotTableProcessor(BadWB2()).extract_all_pivot_tables()  # type: ignore[arg-type]

    # missing-sheet branches in pivot methods
    with pytest.raises(PivotTableError):
        pivots.modify_pivot_table("Missing", 0, {})
    with pytest.raises(PivotTableError):
        pivots.delete_pivot_table("Missing", 0)
    with pytest.raises(PivotTableError):
        pivots.get_pivot_data_summary("Missing", 0)

    # _get_pivot_style branch when attribute missing
    assert pivots._get_pivot_style(object()) == {}

    # exporters exception branches + df.empty stats
    proc = ExcelProcessor()
    proc.load_workbook(str(xlsx), read_only=False)
    exporter = DataExporter(proc)

    monkeypatch.setattr(proc, "get_worksheet_data", lambda *_a, **_k: (_ for _ in ()).throw(RuntimeError("boom")))
    with pytest.raises(WorksheetError):
        exporter.export_worksheet_to_csv("Sales")

    monkeypatch.setattr(proc, "get_workbook_info", lambda *_a, **_k: (_ for _ in ()).throw(RuntimeError("boom")))
    with pytest.raises(WorksheetError):
        exporter.export_workbook_to_json()

    # Empty sheet stats path
    wb2 = openpyxl.Workbook()
    wb2.active.title = "Empty"
    proc_empty = ExcelProcessor()
    proc_empty._workbook = wb2
    exporter2 = DataExporter(proc_empty)
    stats = exporter2.get_summary_statistics("Empty")
    assert stats["total_rows"] == 0

    wb.close()


def test_validation_bool_direct_path():
    assert validate_bool_param(True, "flag") is True


def test_exporters_and_safety_more_branches(tmp_path: Path, monkeypatch):
    xlsx = tmp_path / "book.xlsx"
    _create_workbook(xlsx)

    proc = ExcelProcessor()
    proc.load_workbook(str(xlsx), read_only=False)

    exporter = DataExporter(proc)

    # CSV output_path branch
    csv_path = tmp_path / "out.csv"
    csv_res = exporter.export_worksheet_to_csv("Sales", output_path=str(csv_path))
    assert Path(csv_res["saved_to"]).exists()

    # JSON output_path branch
    json_path = tmp_path / "out.json"
    json_res = exporter.export_workbook_to_json(output_path=str(json_path))
    assert Path(json_res["saved_to"]).exists()

    # pandas exception branch
    def boom(*_a, **_k):
        raise RuntimeError("boom")

    monkeypatch.setattr(proc, "get_worksheet_data", boom)
    with pytest.raises(WorksheetError):
        exporter.export_sheet_to_pandas("Sales")

    # stats exception branch
    monkeypatch.setattr(exporter, "export_sheet_to_pandas", boom)
    with pytest.raises(WorksheetError):
        exporter.get_summary_statistics("Sales")

    # Safety: FileOperationContext restore path
    with pytest.raises(RuntimeError):
        with FileOperationContext(str(xlsx), create_backup=True):
            raise RuntimeError("fail")

    # Safety: cleanup_backup warning branch
    backup_path = xlsx.with_suffix(xlsx.suffix + ".backup")
    backup_path.write_text("x")
    monkeypatch.setattr(type(backup_path), "unlink", lambda *_a, **_k: (_ for _ in ()).throw(RuntimeError("no")))
    cleanup_backup(backup_path)

    proc.close_workbook()
