"""Behavior tests for xlsx_reader.server MCP handlers and dispatch.

Exercises:
- list_tools / list_resources / read_resource (every URI plus unknown)
- call_tool dispatch for every registered tool name
- The defensive paths in call_tool (unknown tool, non-dict args, raised
  exception, non-envelope normalization)
- Each tool's _* implementation function: validation branches, exception
  catching, and success paths against a real on-disk workbook fixture
- run()'s stdio happy path, exception re-raise, and the finally-block
  cleanup that always closes the workbook
"""

import json
from pathlib import Path
from unittest.mock import AsyncMock, patch

import openpyxl
import pytest
from mcp.types import TextContent

from xlsx_reader.server import (
    _add_worksheet,
    _delete_worksheet,
    _export_to_csv,
    _read_workbook_info,
    _read_worksheet_data,
    _save_workbook,
    _update_cell_range,
    _update_cell_value,
    excel_processor,
    handle_call_tool,
    handle_list_resources,
    handle_list_tools,
    handle_read_resource,
    run,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


@pytest.fixture
def sample_workbook(tmp_path):
    """Create a real .xlsx file with one sheet and a small data block."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws["A1"] = "Name"
    ws["B1"] = "Value"
    ws["A2"] = "alpha"
    ws["B2"] = 1
    ws["A3"] = "beta"
    ws["B3"] = 2
    path = tmp_path / "sample.xlsx"
    wb.save(path)
    wb.close()
    yield path
    # Best-effort: ensure the global processor releases the workbook
    # between tests so subsequent tests can reload from disk cleanly.
    try:
        excel_processor.close_workbook()
    except Exception:  # noqa: BLE001
        pass


def _envelope(result):
    """Decode the JSON envelope embedded in a TextContent response."""
    assert isinstance(result, list) and len(result) == 1
    assert isinstance(result[0], TextContent)
    return json.loads(result[0].text)


# ---------------------------------------------------------------------------
# Tool listing
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_list_tools_returns_all_eight_xlsx_tools():
    """handle_list_tools surfaces every registered xlsx-reader tool."""
    tools = await handle_list_tools()
    names = {t.name for t in tools}
    assert names == {
        "read_workbook_info",
        "read_worksheet_data",
        "update_cell_value",
        "update_cell_range",
        "add_worksheet",
        "delete_worksheet",
        "export_to_csv",
        "save_workbook",
    }
    for tool in tools:
        assert tool.description
        assert tool.inputSchema


# ---------------------------------------------------------------------------
# Resources
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_list_resources_returns_two_known_uris():
    """list_resources surfaces supported-formats and server-status."""
    resources = await handle_list_resources()
    uris = {str(r.uri) for r in resources}
    assert uris == {"xlsx://supported-formats", "xlsx://server-status"}


@pytest.mark.asyncio
async def test_read_resource_supported_formats_includes_xlsx_extensions():
    """supported-formats resource lists the four expected file extensions."""
    payload = json.loads(await handle_read_resource("xlsx://supported-formats"))
    assert ".xlsx" in payload["supported_extensions"]
    assert ".xlsm" in payload["supported_extensions"]
    assert payload["max_file_size_mb"] == 200


@pytest.mark.asyncio
async def test_read_resource_server_status_with_no_workbook_loaded():
    """server-status reports workbook_loaded == False when none is loaded."""
    excel_processor.close_workbook()
    payload = json.loads(await handle_read_resource("xlsx://server-status"))
    assert payload["server"] == "xlsx-reader"
    assert payload["workbook_loaded"] is False
    assert payload["current_workbook"] is None


@pytest.mark.asyncio
async def test_read_resource_server_status_with_workbook_loaded(sample_workbook):
    """server-status reports workbook_loaded == True after loading a file."""
    excel_processor.load_workbook(str(sample_workbook), read_only=True)
    payload = json.loads(await handle_read_resource("xlsx://server-status"))
    assert payload["workbook_loaded"] is True
    assert payload["current_workbook"] is not None


@pytest.mark.asyncio
async def test_read_resource_server_status_handles_processor_exception():
    """If get_workbook_info raises, server-status falls back to None workbook info."""
    with patch.object(excel_processor, "is_workbook_loaded", return_value=True), \
         patch.object(
             excel_processor,
             "get_workbook_info",
             side_effect=RuntimeError("processor exploded"),
         ):
        payload = json.loads(await handle_read_resource("xlsx://server-status"))
    assert payload["workbook_loaded"] is False
    assert payload["current_workbook"] is None


@pytest.mark.asyncio
async def test_read_resource_unknown_uri_raises_value_error():
    """Unknown resource URIs raise ValueError with the offending URI."""
    with pytest.raises(ValueError, match="Unknown resource URI"):
        await handle_read_resource("xlsx://does-not-exist")


# ---------------------------------------------------------------------------
# call_tool dispatch — one branch per registered tool
# ---------------------------------------------------------------------------


@pytest.mark.parametrize(
    "tool_name, target_attr",
    [
        ("read_workbook_info", "_read_workbook_info"),
        ("read_worksheet_data", "_read_worksheet_data"),
        ("update_cell_value", "_update_cell_value"),
        ("update_cell_range", "_update_cell_range"),
        ("add_worksheet", "_add_worksheet"),
        ("delete_worksheet", "_delete_worksheet"),
        ("export_to_csv", "_export_to_csv"),
        ("save_workbook", "_save_workbook"),
    ],
)
@pytest.mark.asyncio
async def test_call_tool_routes_to_named_implementation(tool_name, target_attr):
    """Each registered tool name routes through call_tool to its impl."""
    sentinel = {"ok": True, "data": {"routed": tool_name}}
    with patch(
        f"xlsx_reader.server.{target_attr}",
        new=AsyncMock(return_value=sentinel),
    ) as handler:
        result = await handle_call_tool(tool_name, {"file_path": "x.xlsx"})
    handler.assert_awaited_once()
    assert _envelope(result) == sentinel


@pytest.mark.asyncio
async def test_call_tool_unknown_name_returns_user_input_error():
    """Unknown tool names return a UserInput error envelope."""
    payload = _envelope(await handle_call_tool("nonexistent_tool", {}))
    assert payload["ok"] is False
    assert payload["code"] == "UserInput"
    assert "Unknown tool" in payload["message"]


@pytest.mark.asyncio
async def test_call_tool_non_dict_arguments_coerced_to_empty_dict():
    """Non-dict arguments are coerced to {} rather than raising."""
    payload = _envelope(await handle_call_tool("nonexistent_tool", "not-a-dict"))  # type: ignore[arg-type]
    assert payload["ok"] is False
    assert payload["code"] == "UserInput"


@pytest.mark.asyncio
async def test_call_tool_unexpected_exception_returns_internal_error_envelope():
    """An unhandled exception is caught and reported as Internal."""
    with patch(
        "xlsx_reader.server._read_workbook_info",
        new=AsyncMock(side_effect=RuntimeError("boom")),
    ):
        payload = _envelope(await handle_call_tool("read_workbook_info", {}))
    assert payload["ok"] is False
    assert payload["code"] == "Internal"
    assert "boom" in payload["detail"]


@pytest.mark.asyncio
async def test_call_tool_non_envelope_result_normalized_to_envelope():
    """When a tool returns a bare value, server wraps it as {ok, data}."""
    with patch(
        "xlsx_reader.server._read_workbook_info",
        new=AsyncMock(return_value=["a", "b"]),
    ):
        payload = _envelope(await handle_call_tool("read_workbook_info", {}))
    assert payload["ok"] is True
    assert payload["data"] == ["a", "b"]


# ---------------------------------------------------------------------------
# Per-tool implementation: validation branches
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_read_workbook_info_requires_file_path():
    """_read_workbook_info refuses an empty argument set with UserInput."""
    result = await _read_workbook_info({})
    assert result["ok"] is False
    assert result["code"] == "UserInput"


@pytest.mark.asyncio
async def test_read_worksheet_data_requires_file_path():
    """_read_worksheet_data refuses an empty argument set with UserInput."""
    result = await _read_worksheet_data({})
    assert result["ok"] is False
    assert result["code"] == "UserInput"


@pytest.mark.asyncio
async def test_update_cell_value_requires_three_args():
    """_update_cell_value enforces file_path + sheet_name + cell_ref."""
    result = await _update_cell_value({"file_path": "x.xlsx"})
    assert result["ok"] is False
    assert result["code"] == "UserInput"


@pytest.mark.asyncio
async def test_update_cell_value_requires_value_or_formula():
    """_update_cell_value rejects when neither value nor formula provided."""
    result = await _update_cell_value(
        {"file_path": "x.xlsx", "sheet_name": "S", "cell_ref": "A1"}
    )
    assert result["ok"] is False
    assert "value" in result["message"]


@pytest.mark.asyncio
async def test_update_cell_range_requires_full_arg_set():
    """_update_cell_range enforces all four required parameters."""
    result = await _update_cell_range({"file_path": "x.xlsx"})
    assert result["ok"] is False
    assert result["code"] == "UserInput"


@pytest.mark.asyncio
async def test_update_cell_range_requires_values_to_be_a_list():
    """_update_cell_range rejects values when not a list."""
    result = await _update_cell_range(
        {
            "file_path": "x.xlsx",
            "sheet_name": "S",
            "cell_range": "A1:B2",
            "values": "not-a-list",
        }
    )
    assert result["ok"] is False
    assert "2D array" in result["message"]


@pytest.mark.asyncio
async def test_add_worksheet_requires_file_and_sheet_name():
    """_add_worksheet enforces both required parameters."""
    result = await _add_worksheet({"file_path": "x.xlsx"})
    assert result["ok"] is False
    assert result["code"] == "UserInput"


@pytest.mark.asyncio
async def test_delete_worksheet_requires_file_and_sheet_name():
    """_delete_worksheet enforces both required parameters."""
    result = await _delete_worksheet({"file_path": "x.xlsx"})
    assert result["ok"] is False
    assert result["code"] == "UserInput"


@pytest.mark.asyncio
async def test_export_to_csv_requires_file_and_sheet_name():
    """_export_to_csv enforces both required parameters."""
    result = await _export_to_csv({"file_path": "x.xlsx"})
    assert result["ok"] is False
    assert result["code"] == "UserInput"


@pytest.mark.asyncio
async def test_save_workbook_requires_file_path():
    """_save_workbook enforces file_path."""
    result = await _save_workbook({})
    assert result["ok"] is False
    assert result["code"] == "UserInput"


@pytest.mark.asyncio
async def test_save_workbook_rejects_when_no_workbook_loaded():
    """_save_workbook returns UserInput when no workbook is currently loaded."""
    excel_processor.close_workbook()
    result = await _save_workbook({"file_path": "x.xlsx"})
    assert result["ok"] is False
    assert "No workbook" in result["message"]


# ---------------------------------------------------------------------------
# Per-tool implementation: exception-catching branches
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_read_workbook_info_wraps_processor_exception():
    """_read_workbook_info returns Internal envelope when processor raises."""
    with patch.object(
        excel_processor,
        "load_workbook",
        side_effect=RuntimeError("disk error"),
    ):
        result = await _read_workbook_info({"file_path": "x.xlsx"})
    assert result["ok"] is False
    assert result["code"] == "Internal"
    assert "disk error" in result["detail"]


@pytest.mark.asyncio
async def test_read_worksheet_data_wraps_processor_exception():
    """_read_worksheet_data returns Internal envelope when processor raises."""
    with patch.object(
        excel_processor,
        "load_workbook",
        side_effect=RuntimeError("disk error"),
    ):
        result = await _read_worksheet_data({"file_path": "x.xlsx"})
    assert result["ok"] is False
    assert result["code"] == "Internal"


@pytest.mark.asyncio
async def test_update_cell_value_wraps_processor_exception():
    """_update_cell_value returns Internal envelope when processor raises."""
    with patch.object(
        excel_processor,
        "load_workbook",
        side_effect=RuntimeError("disk error"),
    ), patch("xlsx_reader.server.FileOperationContext"):
        result = await _update_cell_value(
            {
                "file_path": "x.xlsx",
                "sheet_name": "S",
                "cell_ref": "A1",
                "value": "v",
            }
        )
    assert result["ok"] is False
    assert result["code"] == "Internal"


@pytest.mark.asyncio
async def test_update_cell_range_wraps_processor_exception():
    """_update_cell_range returns Internal envelope when processor raises."""
    with patch.object(
        excel_processor,
        "load_workbook",
        side_effect=RuntimeError("disk error"),
    ), patch("xlsx_reader.server.FileOperationContext"):
        result = await _update_cell_range(
            {
                "file_path": "x.xlsx",
                "sheet_name": "S",
                "cell_range": "A1:B2",
                "values": [[1, 2], [3, 4]],
            }
        )
    assert result["ok"] is False
    assert result["code"] == "Internal"


@pytest.mark.asyncio
async def test_add_worksheet_wraps_processor_exception():
    """_add_worksheet returns Internal envelope when processor raises."""
    with patch.object(
        excel_processor,
        "load_workbook",
        side_effect=RuntimeError("disk error"),
    ), patch("xlsx_reader.server.FileOperationContext"):
        result = await _add_worksheet(
            {"file_path": "x.xlsx", "sheet_name": "Sheet2"}
        )
    assert result["ok"] is False
    assert result["code"] == "Internal"


@pytest.mark.asyncio
async def test_delete_worksheet_wraps_processor_exception():
    """_delete_worksheet returns Internal envelope when processor raises."""
    with patch.object(
        excel_processor,
        "load_workbook",
        side_effect=RuntimeError("disk error"),
    ), patch("xlsx_reader.server.FileOperationContext"):
        result = await _delete_worksheet(
            {"file_path": "x.xlsx", "sheet_name": "Sheet2"}
        )
    assert result["ok"] is False
    assert result["code"] == "Internal"


@pytest.mark.asyncio
async def test_export_to_csv_wraps_processor_exception():
    """_export_to_csv returns Internal envelope when processor raises."""
    with patch.object(
        excel_processor,
        "load_workbook",
        side_effect=RuntimeError("disk error"),
    ):
        result = await _export_to_csv(
            {"file_path": "x.xlsx", "sheet_name": "Sheet1"}
        )
    assert result["ok"] is False
    assert result["code"] == "Internal"


@pytest.mark.asyncio
async def test_save_workbook_wraps_processor_exception(sample_workbook):
    """_save_workbook returns Internal envelope when processor raises on save."""
    excel_processor.load_workbook(str(sample_workbook), read_only=False)
    with patch.object(
        excel_processor,
        "save_workbook",
        side_effect=RuntimeError("disk error"),
    ):
        result = await _save_workbook({"file_path": str(sample_workbook)})
    assert result["ok"] is False
    assert result["code"] == "Internal"


# ---------------------------------------------------------------------------
# Per-tool implementation: success paths against a real workbook
# ---------------------------------------------------------------------------


@pytest.mark.asyncio
async def test_read_workbook_info_success(sample_workbook):
    """_read_workbook_info returns success envelope for a valid xlsx file."""
    result = await _read_workbook_info({"file_path": str(sample_workbook)})
    assert result["ok"] is True
    assert "Sheet1" in result["data"]["sheet_names"]


@pytest.mark.asyncio
async def test_read_worksheet_data_success(sample_workbook):
    """_read_worksheet_data returns success envelope for an existing sheet."""
    result = await _read_worksheet_data(
        {"file_path": str(sample_workbook), "sheet_name": "Sheet1"}
    )
    assert result["ok"] is True


@pytest.mark.asyncio
async def test_export_to_csv_inline_payload(sample_workbook):
    """_export_to_csv returns CSV inline when no output_path is provided."""
    result = await _export_to_csv(
        {"file_path": str(sample_workbook), "sheet_name": "Sheet1"}
    )
    assert result["ok"] is True
    assert "csv_data" in result["data"]
    assert "Name" in result["data"]["csv_data"]


@pytest.mark.asyncio
async def test_export_to_csv_writes_file_when_output_path_given(
    sample_workbook, tmp_path
):
    """_export_to_csv writes to disk and reports saved_to/file_size."""
    out = tmp_path / "out.csv"
    result = await _export_to_csv(
        {
            "file_path": str(sample_workbook),
            "sheet_name": "Sheet1",
            "output_path": str(out),
        }
    )
    assert result["ok"] is True
    assert result["data"]["saved_to"] == str(out)
    assert out.exists()
    assert "Name" in out.read_text(encoding="utf-8")


@pytest.mark.asyncio
async def test_export_to_csv_drops_header_row_when_disabled(sample_workbook):
    """_export_to_csv strips the first row when include_headers is False."""
    result = await _export_to_csv(
        {
            "file_path": str(sample_workbook),
            "sheet_name": "Sheet1",
            "include_headers": False,
        }
    )
    assert result["ok"] is True
    # Header row should be absent.
    assert "Name" not in result["data"]["csv_data"]


# ---------------------------------------------------------------------------
# run() — server entry point
# ---------------------------------------------------------------------------


class _FakeStdioCtx:
    """Async context manager that yields a fake (read_stream, write_stream)."""

    async def __aenter__(self):
        """Return a fake pair of streams."""
        return (object(), object())

    async def __aexit__(self, exc_type, exc_val, exc_tb):
        """Suppress nothing."""
        return False


@pytest.mark.asyncio
async def test_run_drives_server_run_inside_stdio_context():
    """run() enters stdio_server and awaits server.run exactly once."""
    with patch("mcp.server.stdio.stdio_server", return_value=_FakeStdioCtx()), \
         patch("xlsx_reader.server.server.run", new=AsyncMock()) as mock_run, \
         patch.object(excel_processor, "close_workbook") as cleanup:
        await run()
    mock_run.assert_awaited_once()
    cleanup.assert_called_once()


@pytest.mark.asyncio
async def test_run_reraises_unexpected_exception_and_still_cleans_up():
    """run() re-raises non-cleanup exceptions but still closes the workbook."""
    with patch("mcp.server.stdio.stdio_server", return_value=_FakeStdioCtx()), \
         patch(
             "xlsx_reader.server.server.run",
             new=AsyncMock(side_effect=RuntimeError("network-down")),
         ), patch.object(excel_processor, "close_workbook") as cleanup:
        with pytest.raises(RuntimeError, match="network-down"):
            await run()
    cleanup.assert_called_once()
