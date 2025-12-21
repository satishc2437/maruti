"""Chart processing functionality for Excel workbooks.

Handles chart extraction, modification, and creation.
"""

import logging
from typing import Any, Dict, List, Optional

from ..errors import ChartError, WorksheetError

logger = logging.getLogger(__name__)


class ChartProcessor:
    """Handles Excel chart operations."""

    def __init__(self, workbook):
        """Create a chart processor for an openpyxl workbook."""
        self.workbook = workbook

    def extract_charts_from_sheet(self, sheet_name: str) -> List[Dict[str, Any]]:
        """Extract chart metadata from a worksheet.

        Args:
            sheet_name: Name of worksheet to analyze

        Returns:
            List of chart definitions and properties

        Raises:
            ChartError: If chart extraction fails
        """
        try:
            if sheet_name not in self.workbook.sheetnames:
                raise ChartError(f"Sheet '{sheet_name}' not found")

            sheet = self.workbook[sheet_name]
            charts = []

            # Extract chart objects from worksheet
            if hasattr(sheet, "_charts") and sheet._charts:
                for i, chart in enumerate(sheet._charts):
                    chart_info = {
                        "index": i,
                        "title": getattr(chart, "title", None),
                        "chart_type": self._get_chart_type(chart),
                        "position": self._get_chart_position(chart),
                        "data_range": self._get_chart_data_range(chart),
                        "series": self._extract_chart_series(chart),
                        "style": self._get_chart_style(chart),
                    }
                    charts.append(chart_info)

            return charts

        except Exception as e:
            raise ChartError(f"Failed to extract charts from sheet '{sheet_name}': {e}")

    def extract_all_charts(self) -> Dict[str, List[Dict[str, Any]]]:
        """Extract charts from all worksheets in the workbook.

        Returns:
            Dictionary mapping sheet names to their chart lists

        Raises:
            ChartError: If extraction fails
        """
        try:
            all_charts = {}

            for sheet_name in self.workbook.sheetnames:
                try:
                    charts = self.extract_charts_from_sheet(sheet_name)
                    if charts:
                        all_charts[sheet_name] = charts
                except Exception as e:
                    logger.warning(f"Failed to extract charts from '{sheet_name}': {e}")
                    all_charts[sheet_name] = {"error": str(e)}

            return all_charts

        except Exception as e:
            raise ChartError(f"Failed to extract charts from workbook: {e}")

    def create_chart(
        self,
        sheet_name: str,
        chart_type: str,
        data_range: str,
        title: Optional[str] = None,
        position: Optional[Dict[str, int]] = None,
    ) -> Dict[str, Any]:
        """Create a new chart in the specified worksheet.

        Args:
            sheet_name: Name of worksheet to add chart to
            chart_type: Type of chart (e.g., 'column', 'line', 'pie')
            data_range: Range of cells containing chart data
            title: Chart title
            position: Chart position {'col': x, 'row': y}

        Returns:
            Information about the created chart

        Raises:
            ChartError: If chart creation fails
        """
        try:
            from openpyxl.chart import (
                AreaChart,
                BarChart,
                LineChart,
                PieChart,
                Reference,
                ScatterChart,
            )

            if sheet_name not in self.workbook.sheetnames:
                raise ChartError(f"Sheet '{sheet_name}' not found")

            sheet = self.workbook[sheet_name]

            # Create chart based on type
            chart_classes = {
                "column": BarChart,
                "bar": BarChart,
                "line": LineChart,
                "pie": PieChart,
                "area": AreaChart,
                "scatter": ScatterChart,
            }

            if chart_type.lower() not in chart_classes:
                raise ChartError(f"Unsupported chart type: {chart_type}")

            chart_class = chart_classes[chart_type.lower()]
            chart = chart_class()

            # Set chart properties
            if title:
                chart.title = title

            # Parse data range and create reference
            data_ref = Reference(sheet, range_string=data_range)
            chart.add_data(data_ref, titles_from_data=True)

            # Set position
            if position:
                chart.anchor = f"{self._col_to_letter(position.get('col', 1))}{position.get('row', 1)}"
            else:
                chart.anchor = "E5"  # Default position

            # Add chart to worksheet
            sheet.add_chart(chart)

            return {
                "sheet_name": sheet_name,
                "chart_type": chart_type,
                "title": title,
                "data_range": data_range,
                "position": position or {"col": 5, "row": 5},
                "created": True,
            }

        except Exception as e:
            raise ChartError(f"Failed to create chart: {e}")

    def modify_chart(
        self, sheet_name: str, chart_index: int, modifications: Dict[str, Any]
    ) -> Dict[str, Any]:
        """Modify an existing chart in the worksheet.

        Args:
            sheet_name: Name of worksheet containing the chart
            chart_index: Index of chart in the worksheet
            modifications: Dictionary of properties to modify

        Returns:
            Information about the modified chart

        Raises:
            ChartError: If chart modification fails
        """
        try:
            if sheet_name not in self.workbook.sheetnames:
                raise ChartError(f"Sheet '{sheet_name}' not found")

            sheet = self.workbook[sheet_name]

            if not hasattr(sheet, "_charts") or not sheet._charts:
                raise ChartError(f"No charts found in sheet '{sheet_name}'")

            if chart_index >= len(sheet._charts):
                raise ChartError(f"Chart index {chart_index} out of range")

            chart = sheet._charts[chart_index]
            applied_modifications = {}

            # Apply modifications
            for prop, value in modifications.items():
                try:
                    if prop == "title":
                        chart.title = value
                        applied_modifications["title"] = value
                    elif prop == "data_range":
                        # Update data range
                        from openpyxl.chart import Reference

                        data_ref = Reference(sheet, range_string=value)
                        chart.series.clear()
                        chart.add_data(data_ref, titles_from_data=True)
                        applied_modifications["data_range"] = value
                    elif prop == "position":
                        col = value.get("col", 1)
                        row = value.get("row", 1)
                        chart.anchor = f"{self._col_to_letter(col)}{row}"
                        applied_modifications["position"] = value
                    else:
                        logger.warning(f"Unknown chart property: {prop}")

                except Exception as e:
                    logger.warning(f"Failed to apply modification {prop}: {e}")

            return {
                "sheet_name": sheet_name,
                "chart_index": chart_index,
                "applied_modifications": applied_modifications,
                "modified": True,
            }

        except Exception as e:
            raise ChartError(f"Failed to modify chart: {e}")

    def delete_chart(self, sheet_name: str, chart_index: int) -> Dict[str, Any]:
        """Delete a chart from the worksheet.

        Args:
            sheet_name: Name of worksheet containing the chart
            chart_index: Index of chart to delete

        Returns:
            Information about the deleted chart

        Raises:
            ChartError: If chart deletion fails
        """
        try:
            if sheet_name not in self.workbook.sheetnames:
                raise ChartError(f"Sheet '{sheet_name}' not found")

            sheet = self.workbook[sheet_name]

            if not hasattr(sheet, "_charts") or not sheet._charts:
                raise ChartError(f"No charts found in sheet '{sheet_name}'")

            if chart_index >= len(sheet._charts):
                raise ChartError(f"Chart index {chart_index} out of range")

            # Remove chart
            chart = sheet._charts.pop(chart_index)

            return {
                "sheet_name": sheet_name,
                "chart_index": chart_index,
                "remaining_charts": len(sheet._charts),
                "deleted": True,
            }

        except Exception as e:
            raise ChartError(f"Failed to delete chart: {e}")

    # Helper methods

    def _get_chart_type(self, chart) -> str:
        """Get the type of chart."""
        chart_type_map = {
            "barChart": "column",
            "lineChart": "line",
            "pieChart": "pie",
            "areaChart": "area",
            "scatterChart": "scatter",
        }

        chart_class = chart.__class__.__name__.lower()
        for key, value in chart_type_map.items():
            if key.lower() in chart_class:
                return value

        return chart_class

    def _get_chart_position(self, chart) -> Dict[str, Any]:
        """Get chart position information."""
        try:
            if hasattr(chart, "anchor"):
                anchor = chart.anchor
                return {
                    "anchor": str(anchor),
                    "col": getattr(anchor, "col", None),
                    "row": getattr(anchor, "row", None),
                }
        except Exception:
            pass

        return {"anchor": None, "col": None, "row": None}

    def _get_chart_data_range(self, chart) -> Optional[str]:
        """Get the data range used by the chart."""
        try:
            if hasattr(chart, "series") and chart.series:
                # Get first series data range
                first_series = chart.series[0]
                if hasattr(first_series, "values") and first_series.values:
                    return str(first_series.values)
        except Exception:
            pass

        return None

    def _extract_chart_series(self, chart) -> List[Dict[str, Any]]:
        """Extract information about chart data series."""
        series_info = []

        try:
            if hasattr(chart, "series") and chart.series:
                for i, series in enumerate(chart.series):
                    info = {
                        "index": i,
                        "title": getattr(series, "title", None),
                        "values_range": str(getattr(series, "values", None)),
                        "categories_range": str(getattr(series, "cat", None)),
                    }
                    series_info.append(info)
        except Exception as e:
            logger.warning(f"Failed to extract series info: {e}")

        return series_info

    def _get_chart_style(self, chart) -> Dict[str, Any]:
        """Get chart styling information."""
        style_info = {}

        try:
            # Extract basic style properties
            if hasattr(chart, "style"):
                style_info["style_id"] = chart.style
            if hasattr(chart, "legend"):
                style_info["legend"] = {
                    "position": getattr(chart.legend, "position", None)
                    if chart.legend
                    else None
                }
        except Exception as e:
            logger.warning(f"Failed to extract chart style: {e}")

        return style_info

    def _col_to_letter(self, col_num: int) -> str:
        """Convert column number to Excel letter format."""
        result = ""
        while col_num > 0:
            col_num -= 1
            result = chr(65 + col_num % 26) + result
            col_num //= 26
        return result
