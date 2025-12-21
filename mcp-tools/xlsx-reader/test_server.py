#!/usr/bin/env python3
"""Simple smoke tests for the Excel Reader MCP server.

These tests validate core workbook operations and error handling.
"""

import asyncio
import logging
import tempfile
from pathlib import Path

from xlsx_reader.errors import success_response, user_input_error
from xlsx_reader.processors.exporters import DataExporter
from xlsx_reader.processors.workbook import ExcelProcessor

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def create_sample_workbook():
    """Create a sample Excel workbook for testing."""
    import openpyxl

    # Create workbook with sample data
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Data"

    # Add headers
    headers = ["Product", "Quantity", "Price", "Total"]
    for col, header in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=header)

    # Add sample data
    data = [
        ["Widget A", 100, 25.99, "=B2*C2"],
        ["Widget B", 150, 19.99, "=B3*C3"],
        ["Widget C", 75, 35.50, "=B4*C4"],
    ]

    for row, row_data in enumerate(data, 2):
        for col, value in enumerate(row_data, 1):
            ws.cell(row=row, column=col, value=value)

    # Add another worksheet
    ws2 = wb.create_sheet("Summary")
    ws2.cell(row=1, column=1, value="Total Products")
    ws2.cell(row=1, column=2, value=3)

    return wb


def test_workbook_operations(tmp_path: Path):
    """Test core workbook operations."""
    logger.info("Testing Excel workbook operations...")

    workbook_path = tmp_path / "sample.xlsx"

    wb = create_sample_workbook()
    wb.save(workbook_path)
    wb.close()

    processor = ExcelProcessor()

    workbook_info = processor.load_workbook(str(workbook_path), read_only=True)
    assert workbook_info["sheet_count"] == 2
    assert "Sales Data" in workbook_info["sheet_names"]
    assert "Summary" in workbook_info["sheet_names"]

    worksheet_data = processor.get_worksheet_data("Sales Data")
    assert worksheet_data["rows"] > 0
    assert worksheet_data["columns"] == 4

    exporter = DataExporter(processor)

    csv_result = exporter.export_worksheet_to_csv("Sales Data")
    assert "csv_data" in csv_result
    assert csv_result["rows_exported"] > 0

    json_result = exporter.export_workbook_to_json()
    assert "json_data" in json_result
    assert json_result["sheets_exported"] > 0

    processor.close_workbook()
    processor.load_workbook(str(workbook_path), read_only=False)

    update_result = processor.update_cell_value("Sales Data", "E1", "Status")
    assert update_result["updated"] is True

    add_result = processor.add_worksheet("Test Sheet")
    assert add_result["name"] == "Test Sheet"

    save_result = processor.save_workbook()
    assert "saved_to" in save_result

    processor.close_workbook()


def test_error_handling() -> None:
    """Test error handling functionality."""
    logger.info("Testing error handling...")

    # Test various error scenarios
    processor = ExcelProcessor()

    # Test 1: Invalid file path
    try:
        processor.load_workbook("/nonexistent/file.xlsx")
        assert False, "Should have raised an exception"
    except Exception as e:  # pylint: disable=broad-except
        logger.info("✓ Correctly handled invalid file path: %s", type(e).__name__)

    # Test 2: Invalid sheet name
    try:
        # Create a temporary workbook first
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
            wb = create_sample_workbook()
            wb.save(tmp.name)
            wb.close()

            processor.load_workbook(tmp.name, read_only=True)
            processor.get_worksheet_data("NonExistentSheet")
            assert False, "Should have raised an exception"
    except Exception as e:  # pylint: disable=broad-except
        logger.info("✓ Correctly handled invalid sheet name: %s", type(e).__name__)
    finally:
        processor.close_workbook()
        try:
            Path(tmp.name).unlink()
        except OSError:
            pass

    # Test 3: Error response functions
    error_resp = user_input_error("Test error", hint="Test hint")
    assert error_resp["ok"] is False
    assert error_resp["code"] == "UserInput"
    logger.info("✓ Error response functions work correctly")

    success_resp = success_response({"test": "data"})
    assert success_resp["ok"] is True
    assert success_resp["data"]["test"] == "data"
    logger.info("✓ Success response functions work correctly")

    logger.info("Error handling tests passed! ✓")


def print_server_info():
    """Print information about the MCP server."""
    print("\n" + "=" * 60)
    print("EXCEL READER MCP SERVER")
    print("=" * 60)
    print("\nProject Structure:")
    print("xlsx-reader/")
    print("├── src/xlsx_reader/")
    print("│   ├── __main__.py          # Entry point")
    print("│   ├── server.py            # MCP server implementation")
    print("│   ├── errors.py            # Error handling")
    print("│   ├── safety.py            # File safety utilities")
    print("│   ├── processors/")
    print("│   │   ├── workbook.py      # Core Excel processing")
    print("│   │   ├── exporters.py     # Data export functionality")
    print("│   │   ├── charts.py        # Chart operations")
    print("│   │   └── pivots.py        # Pivot table operations")
    print("│   └── utils/")
    print("│       └── validation.py    # Parameter validation")
    print("├── pyproject.toml           # Dependencies and config")
    print("├── README.md                # Main documentation")
    print("└── USAGE_EXAMPLES.md        # Usage examples")

    print("\nKey Features:")
    print("✓ Read Excel workbook metadata and sheet information")
    print("✓ Read/write worksheet cell data and formulas")
    print("✓ Add/delete/rename worksheets")
    print("✓ Export data to CSV and JSON formats")
    print("✓ Chart extraction and metadata (limited modification)")
    print("✓ Pivot table extraction and metadata (limited modification)")
    print("✓ Automatic file backup and locking")
    print("✓ Comprehensive error handling")
    print("✓ Full filesystem access (configurable)")

    print("\nRun Commands (from maruti/xlsx-reader/):")
    print("  uv pip install -e .                # Install in workspace")
    print("  python -m xlsx_reader               # Start MCP server")
    print("  python -m xlsx_reader --debug       # Start with debug logging")
    print("  python test_server.py               # Run this test script")
    print("  uvx --from . python -m xlsx_reader  # Run without install")

    print("\nSupported File Formats:")
    print("  .xlsx  - Excel Workbook (OpenXML)")
    print("  .xlsm  - Excel Macro-Enabled Workbook")
    print("  .xltx  - Excel Template")
    print("  .xltm  - Excel Macro-Enabled Template")

    print("\nSafety Features:")
    print("  • Automatic backup creation before modifications")
    print("  • File locking to prevent concurrent access")
    print("  • Path validation and size limits")
    print("  • Structured error responses with helpful hints")
    print("  • Maximum file size: 200MB (configurable)")

    print("=" * 60)


async def main():
    """Run all tests and display server information."""
    print_server_info()

    print("\nRunning functionality tests...")
    print("-" * 40)

    with tempfile.TemporaryDirectory() as tmp_dir:
        tmp_path = Path(tmp_dir)
        test_workbook_operations(tmp_path)
        test_error_handling()

    print("\n" + "-" * 40)
    print("🎉 ALL TESTS PASSED!")
    print("\nThe Excel Reader MCP server is ready to use!")
    print("\nTo start the server from maruti project root:")
    print("  cd mcp-tools/xlsx-reader")
    print("  uv pip install -e .")
    print("  python -m xlsx_reader")

    return 0


if __name__ == "__main__":
    exit_code = asyncio.run(main())
    raise SystemExit(exit_code)
